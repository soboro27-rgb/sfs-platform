from fastapi import APIRouter, Request, Depends
from fastapi.responses import RedirectResponse, HTMLResponse, StreamingResponse, FileResponse
from sqlalchemy.orm import Session
from database import get_db
import models
from auth import require_admin
from config import templates
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
import os
import base64
from pathlib import Path
from stamp_data import STAMP_B64

router = APIRouter()

IN_PROGRESS_STATUSES = ["approved", "scheduled", "schedule_confirmed", "collected", "priced"]


def _check(request: Request):
    user = require_admin(request)
    if not user:
        return None, RedirectResponse("/login", status_code=302)
    return user, None


def _get_fee_rate(db: Session) -> float:
    config = db.query(models.SystemConfig).filter(models.SystemConfig.key == "welfare_fee_rate").first()
    if config:
        try:
            return float(config.value)
        except (ValueError, TypeError):
            return 0.0
    return 0.0


@router.get("/dashboard", response_class=HTMLResponse)
def dashboard(request: Request, db: Session = Depends(get_db)):
    user, redir = _check(request)
    if redir:
        return redir

    total = db.query(models.Application).count()
    submitted = db.query(models.Application).filter(models.Application.status == "submitted").count()
    in_progress = db.query(models.Application).filter(models.Application.status.in_(IN_PROGRESS_STATUSES)).count()
    completed = db.query(models.Application).filter(
        models.Application.status.in_(["branch_confirmed", "completed"])
    ).count()
    recent = (
        db.query(models.Application)
        .order_by(models.Application.updated_at.desc())
        .limit(10)
        .all()
    )

    return templates.TemplateResponse(
        "admin/dashboard.html",
        {
            "request": request,
            "session": request.session,
            "total": total,
            "submitted": submitted,
            "in_progress": in_progress,
            "completed": completed,
            "recent": recent,
        },
    )


@router.get("/applications", response_class=HTMLResponse)
def application_list(request: Request, status: str = "", db: Session = Depends(get_db)):
    user, redir = _check(request)
    if redir:
        return redir

    query = db.query(models.Application)
    if status:
        query = query.filter(models.Application.status == status)
    applications = query.order_by(models.Application.updated_at.desc()).all()

    return templates.TemplateResponse(
        "admin/application_list.html",
        {
            "request": request,
            "session": request.session,
            "applications": applications,
            "current_status": status,
        },
    )


@router.get("/applications/{app_id}", response_class=HTMLResponse)
def application_detail(request: Request, app_id: int, error: str = "", db: Session = Depends(get_db)):
    user, redir = _check(request)
    if redir:
        return redir

    app = db.query(models.Application).filter(models.Application.id == app_id).first()
    if not app:
        return RedirectResponse("/admin/applications", status_code=302)

    fee_rate = _get_fee_rate(db)
    return templates.TemplateResponse(
        "admin/application_detail.html",
        {"request": request, "session": request.session, "app": app, "welfare_fee_rate": fee_rate, "error": error},
    )


@router.post("/applications/{app_id}/approve")
def approve(request: Request, app_id: int, db: Session = Depends(get_db)):
    user, redir = _check(request)
    if redir:
        return redir
    if user["role"] != "coretail":
        return RedirectResponse(f"/admin/applications/{app_id}", status_code=302)

    app = db.query(models.Application).filter(
        models.Application.id == app_id,
        models.Application.status == "submitted",
    ).first()
    if app:
        app.status = "approved"
        app.approved_at = datetime.now()
        app.updated_at = datetime.now()
        db.commit()

    return RedirectResponse(f"/admin/applications/{app_id}", status_code=302)


@router.post("/applications/{app_id}/schedule")
async def set_schedule(request: Request, app_id: int, db: Session = Depends(get_db)):
    user, redir = _check(request)
    if redir:
        return redir
    if user["role"] != "coretail":
        return RedirectResponse(f"/admin/applications/{app_id}", status_code=302)

    form = await request.form()

    app = db.query(models.Application).filter(
        models.Application.id == app_id,
        models.Application.status == "approved",
    ).first()

    if app:
        if not app.schedule:
            sched = models.VisitSchedule(application_id=app_id)
            db.add(sched)
            db.flush()
            db.refresh(app)

        app.schedule.visit_date = form.get("visit_date", "")
        app.schedule.visit_time = form.get("visit_time", "")
        app.schedule.collector_name = form.get("collector_name", "")
        app.schedule.collector_phone = form.get("collector_phone", "")
        app.schedule.notes = form.get("notes", "")
        app.status = "scheduled"
        app.updated_at = datetime.now()
        db.commit()

    return RedirectResponse(f"/admin/applications/{app_id}", status_code=302)


@router.post("/applications/{app_id}/collect")
async def mark_collected(request: Request, app_id: int, db: Session = Depends(get_db)):
    user, redir = _check(request)
    if redir:
        return redir
    if user["role"] != "coretail":
        return RedirectResponse(f"/admin/applications/{app_id}", status_code=302)

    form = await request.form()

    app = db.query(models.Application).filter(
        models.Application.id == app_id,
        models.Application.status == "schedule_confirmed",
    ).first()
    if not app:
        return RedirectResponse(f"/admin/applications/{app_id}", status_code=302)

    wipe_date = form.get("wipe_date", "").strip()
    technician_company = form.get("technician_company", "").strip()
    technician_name = form.get("technician_name", "").strip()

    if not wipe_date or not technician_company or not technician_name:
        return RedirectResponse(
            f"/admin/applications/{app_id}?error=wipe_required",
            status_code=302,
        )

    if not app.data_wipe_record:
        wipe_rec = models.DataWipeRecord(application_id=app_id)
        db.add(wipe_rec)
        db.flush()
        db.refresh(app)

    wr = app.data_wipe_record
    wr.wipe_date = wipe_date
    wr.technician_company = technician_company
    wr.technician_name = technician_name
    wr.wipe_location = form.get("wipe_location", "").strip()
    wr.blancco_method = form.get("blancco_method", "").strip()
    wr.destruction_method = form.get("destruction_method", "").strip()
    wr.destruction_certificate_no = form.get("destruction_certificate_no", "").strip()
    wr.notes = form.get("wipe_notes", "").strip()
    wr.updated_at = datetime.now()

    # 천공사진 저장
    photo_dir = Path("static/uploads/wipe_photos")
    photo_dir.mkdir(parents=True, exist_ok=True)
    for idx, field in enumerate(["destruction_photo1", "destruction_photo2"], 1):
        photo = form.get(field)
        if photo and getattr(photo, "filename", None):
            ext = Path(photo.filename).suffix.lower()
            if ext not in (".jpg", ".jpeg", ".png", ".gif", ".webp"):
                ext = ".jpg"
            fname = f"{app_id}_photo{idx}{ext}"
            fpath = photo_dir / fname
            contents = await photo.read()
            if contents:
                fpath.write_bytes(contents)
                if idx == 1:
                    wr.destruction_photo1 = str(fpath).replace("\\", "/")
                else:
                    wr.destruction_photo2 = str(fpath).replace("\\", "/")

    # 블랑코 리포트 파일 업로드
    blancco_report_dir = Path("static/uploads/blancco_reports")
    blancco_report_dir.mkdir(parents=True, exist_ok=True)
    blancco_file = form.get("blancco_report_file")
    if blancco_file and getattr(blancco_file, "filename", None):
        ext = Path(blancco_file.filename).suffix.lower()
        if ext not in (".pdf", ".xlsx", ".xls"):
            ext = ".pdf"
        fname = f"{app_id}_blancco_report{ext}"
        fpath = blancco_report_dir / fname
        contents = await blancco_file.read()
        if contents:
            fpath.write_bytes(contents)
            wr.blancco_report_file = str(fpath).replace("\\", "/")

    app.status = "collected"
    app.updated_at = datetime.now()
    db.commit()

    return RedirectResponse(f"/admin/applications/{app_id}", status_code=302)


def _make_thin_border():
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _make_medium_border():
    med = Side(style="medium")
    return Border(left=med, right=med, top=med, bottom=med)


def _set_border(ws, row, col_start, col_end, border):
    for col in range(col_start, col_end + 1):
        ws.cell(row=row, column=col).border = border


# 자산 카테고리 → 인수인계서 품목 매핑
_ASSET_TO_TRANSFER = {
    "PC": "컴퓨터(본체)",
    "노트북": "노트북",
    "태블릿": "태블릿pc",
    "모바일": "태블릿pc",
    "프린터": "프린터",
    "복합기": "기타",
    "기타전산기기": "기타",
}
_TRANSFER_ROWS = [
    "컴퓨터(본체)", "모니터", "노트북", "스캐너", "프린터",
    "워크스테이션", "네트워크장비", "일체형pc", "태블릿pc", "기타",
]
_DISPOSAL_ROWS = ["컴퓨터 본체", "모니터", "노트북", "TV", "기타장비류"]
_DISPOSAL_CAT_MAP = {
    "컴퓨터 본체": ["PC"],
    "모니터": [],
    "노트북": ["노트북"],
    "TV": [],
    "기타장비류": ["태블릿", "모바일", "프린터", "복합기", "기타전산기기"],
}


def _agg_transfer(assets):
    qtys = {k: 0 for k in _TRANSFER_ROWS}
    for a in assets:
        key = _ASSET_TO_TRANSFER.get(a.category, "기타")
        qtys[key] += a.quantity
    return qtys


def _agg_disposal(assets):
    qtys = {k: 0 for k in _DISPOSAL_ROWS}
    for a in assets:
        for label, cats in _DISPOSAL_CAT_MAP.items():
            if a.category in cats:
                qtys[label] += a.quantity
    return qtys


def _build_shredding_sheet(ws, app):
    wr = app.data_wipe_record
    tb = _make_thin_border()
    C = Alignment(horizontal="center", vertical="center", wrap_text=True)
    L = Alignment(horizontal="left", vertical="center")

    for i, w in enumerate([3, 10, 20, 10, 10, 10, 10, 3], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    bar_fill = PatternFill(start_color="B0BEC5", end_color="B0BEC5", fill_type="solid")
    ws.merge_cells("A1:H1")
    ws["A1"].fill = bar_fill
    ws.row_dimensions[1].height = 8

    ws.merge_cells("A3:H3")
    ws["A3"].value = "파  쇄  증  명  서"
    ws["A3"].font = Font(bold=True, size=22)
    ws["A3"].alignment = C
    ws.row_dimensions[3].height = 40

    ws.merge_cells("A4:H4")
    ws["A4"].value = "CERTIFICATE  OF  DESTRUCTION"
    ws["A4"].font = Font(size=11, italic=True, color="555555")
    ws["A4"].alignment = C
    ws.row_dimensions[4].height = 18

    ws.merge_cells("B5:G5")
    ws.row_dimensions[5].height = 3
    for col in range(2, 8):
        ws.cell(row=5, column=col).border = Border(bottom=Side(style="medium"))

    ws.merge_cells("B7:G7")
    ws["B7"].value = "이증명서는 ㈜월드와이드 메모리가 아래와 같이"
    ws["B7"].font = Font(bold=True, size=11)
    ws["B7"].alignment = C
    ws.row_dimensions[7].height = 20

    ws.merge_cells("B8:G8")
    ws["B8"].value = "귀사의 보안 자료를 파쇄 하였음을 증명합니다."
    ws["B8"].font = Font(bold=True, size=11)
    ws["B8"].alignment = C
    ws.row_dimensions[8].height = 20
    ws.row_dimensions[9].height = 12

    total_qty = sum(a.quantity for a in app.assets if a.data_wiped in ("파쇄완료", "파쇄")) or sum(a.quantity for a in app.assets)
    fields = [
        ("고객사 :", app.user.branch_name),
        ("파쇄종류 :", wr.destruction_method if wr else ""),
        ("파쇄수량 :", f"{total_qty}개"),
        ("파쇄장소 :", wr.wipe_location if wr else ""),
        ("파쇄일시 :", wr.wipe_date if wr else ""),
    ]
    lf = PatternFill(start_color="ECEFF1", end_color="ECEFF1", fill_type="solid")
    ul = Border(bottom=Side(style="medium"))
    for idx, (lbl, val) in enumerate(fields, 10):
        r = idx
        ws.cell(row=r, column=2, value=f"  {lbl}").font = Font(bold=True, size=11)
        ws.cell(row=r, column=2).fill = lf
        ws.cell(row=r, column=2).border = ul
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
        ws.cell(row=r, column=3, value=val).font = Font(size=11)
        ws.cell(row=r, column=3).border = ul
        ws.row_dimensions[r].height = 24

    ws.row_dimensions[15].height = 20
    ws.merge_cells("D17:E17")
    ws["D17"].value = "대표이사"
    ws["D17"].font = Font(bold=True, size=13)
    ws["D17"].alignment = C
    ws.merge_cells("F17:G17")
    ws["F17"].value = "최병진 장성대"
    ws["F17"].font = Font(bold=True, size=13)
    ws["F17"].alignment = L
    ws.row_dimensions[17].height = 28

    try:
        stamp_bytes = base64.b64decode(STAMP_B64)
        stamp = XLImage(BytesIO(stamp_bytes))
        stamp.width, stamp.height = 80, 80
        ws.add_image(stamp, "F15")
    except Exception as e:
        print(f"[직인] 삽입 실패: {e}")

    ws.merge_cells("A19:H19")
    ws["A19"].fill = bar_fill
    ws.row_dimensions[19].height = 8

    has_photo = wr and (wr.destruction_photo1 or wr.destruction_photo2)
    if has_photo:
        photo_title_row = 21
        ws.merge_cells(f"A{photo_title_row}:H{photo_title_row}")
        ws[f"A{photo_title_row}"].value = "[ 천 공 사 진 ]"
        ws[f"A{photo_title_row}"].font = Font(bold=True, size=12)
        ws[f"A{photo_title_row}"].alignment = C
        ws.row_dimensions[photo_title_row].height = 22
        for pr in range(22, 36):
            ws.row_dimensions[pr].height = 14
        if wr.destruction_photo1 and os.path.exists(wr.destruction_photo1):
            try:
                img1 = XLImage(wr.destruction_photo1)
                img1.width, img1.height = 220, 165
                ws.add_image(img1, "B22")
            except Exception:
                pass
        if wr.destruction_photo2 and os.path.exists(wr.destruction_photo2):
            try:
                img2 = XLImage(wr.destruction_photo2)
                img2.width, img2.height = 220, 165
                ws.add_image(img2, "E22")
            except Exception:
                pass


def _generate_shredding_excel(app) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "파쇄증명서"
    _build_shredding_sheet(ws, app)
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def _build_transfer_sheet(ws, app):
    wr = app.data_wipe_record
    tb = _make_thin_border()
    C = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, w in enumerate([2, 14, 20, 8, 8, 20, 8, 8, 20, 8, 8, 2], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:L1")
    ws["A1"].value = "물 품  인 계 인 수 서"
    ws["A1"].font = Font(bold=True, size=18)
    ws["A1"].alignment = C
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 10

    def section(ws, start_row, num, title, lines):
        ws.cell(row=start_row, column=2, value=f"{num}. {title}").font = Font(bold=True, size=11)
        ws.row_dimensions[start_row].height = 18
        for i, (lbl, val) in enumerate(lines):
            r = start_row + 1 + i
            ws.cell(row=r, column=2, value=f"   ▶ {lbl}").font = Font(size=10)
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=11)
            c = ws.cell(row=r, column=3, value=val)
            c.font = Font(size=10)
            c.border = Border(bottom=Side(style="thin"))
            ws.row_dimensions[r].height = 17

    section(ws, 3, 1, "인계자 :", [
        ("주   소 :", ""),
        ("전   화 :", ""),
        ("담당자 :", f"{app.user.branch_name}  (서명)"),
    ])
    section(ws, 8, 2, "인수자 :", [
        ("주   소 :", "경기도 고양시 일산동구 장대길 64-47"),
        ("전   화 :", "070-4888-3070"),
        ("담당자 :", "장일운  (서명)"),
    ])

    ws.merge_cells("B13:K13")
    ws["B13"].value = "아래 물품을 정히 인수합니다."
    ws["B13"].font = Font(size=10)
    ws.row_dimensions[13].height = 16

    qtys = _agg_transfer(app.assets)
    t_fill = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")
    t_font = Font(color="FFFFFF", bold=True, size=10)

    for col, val in [(2, "품  명"), (5, "수  량"), (6, "품  명"), (9, "수  량"), (10, "품  명")]:
        c = ws.cell(row=14, column=col, value=val)
        c.font = t_font; c.fill = t_fill; c.alignment = C; c.border = tb
    ws.merge_cells("B14:D14")
    ws.merge_cells("F14:H14")
    ws.merge_cells("J14:K14")
    ws.cell(row=14, column=5).font = t_font
    ws.cell(row=14, column=5).fill = t_fill
    ws.cell(row=14, column=5).alignment = C
    ws.cell(row=14, column=5).border = tb
    ws.cell(row=14, column=9).font = t_font
    ws.cell(row=14, column=9).fill = t_fill
    ws.cell(row=14, column=9).alignment = C
    ws.cell(row=14, column=9).border = tb
    ws.row_dimensions[14].height = 18

    items = _TRANSFER_ROWS
    for idx, name in enumerate(items[:5]):
        r = 15 + idx
        qty = qtys[name]
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        ws.cell(row=r, column=2, value=name).border = tb
        ws.cell(row=r, column=2).alignment = C
        ws.cell(row=r, column=5, value=qty if qty else "").border = tb
        ws.cell(row=r, column=5).alignment = C
        ws.row_dimensions[r].height = 16

    for idx, name in enumerate(items[5:]):
        r = 15 + idx
        qty = qtys[name]
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)
        ws.cell(row=r, column=6, value=name).border = tb
        ws.cell(row=r, column=6).alignment = C
        ws.cell(row=r, column=9, value=qty if qty else "").border = tb
        ws.cell(row=r, column=9).alignment = C

    ws.row_dimensions[22].height = 10
    ws.merge_cells("B23:K23")
    ws["B23"].value = f"수령일자  :  {wr.wipe_date if wr else ''}  (수거일 기준)"
    ws["B23"].font = Font(size=11)
    ws["B23"].alignment = C
    ws.row_dimensions[23].height = 22
    ws.row_dimensions[24].height = 10
    ws.merge_cells("H25:K25")
    ws["H25"].value = "대표이사    최병진 장성대"
    ws["H25"].font = Font(bold=True, size=12)
    ws["H25"].alignment = C
    ws.row_dimensions[25].height = 22

    try:
        stamp_bytes = base64.b64decode(STAMP_B64)
        stamp2 = XLImage(BytesIO(stamp_bytes))
        stamp2.width, stamp2.height = 80, 80
        ws.add_image(stamp2, "K23")
    except Exception as e:
        print(f"[직인] 물품인수인계증 삽입 실패: {e}")


def _generate_transfer_excel(app) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "물품인수인계증"
    _build_transfer_sheet(ws, app)
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def _build_disposal_sheet(ws, app):
    wr = app.data_wipe_record
    tb = _make_thin_border()
    C = Alignment(horizontal="center", vertical="center", wrap_text=True)
    L = Alignment(horizontal="left", vertical="center")

    for i, w in enumerate([2, 16, 22, 14, 14, 14, 2], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:G1")
    ws["A1"].value = "폐  기  확  인  서"
    ws["A1"].font = Font(bold=True, size=20)
    ws["A1"].alignment = C
    ws.row_dimensions[1].height = 38
    ws.row_dimensions[2].height = 10

    lbl_f = PatternFill(start_color="ECEFF1", end_color="ECEFF1", fill_type="solid")
    sec_f = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")
    sec_ft = Font(color="FFFFFF", bold=True, size=10)

    ws.merge_cells(start_row=3, start_column=2, end_row=6, end_column=2)
    ws["B3"].value = "폐기\n의뢰자"
    ws["B3"].alignment = C
    ws["B3"].font = sec_ft
    ws["B3"].fill = sec_f
    ws["B3"].border = tb

    dep_info = [
        ("회사명 :", app.user.branch_name, None, None),
        ("소재지 :", "", None, None),
        ("대표자 :", "", "전자우편주소", ""),
        ("전화번호 :", "", "팩스번호", ""),
    ]
    for idx, (lbl, val, lbl2, val2) in enumerate(dep_info):
        r = 3 + idx
        ws.cell(row=r, column=3, value=f"  {lbl}").font = Font(bold=True, size=10)
        ws.cell(row=r, column=3).fill = lbl_f
        ws.cell(row=r, column=3).border = tb
        if lbl2:
            ws.cell(row=r, column=4, value=val).border = tb
            ws.cell(row=r, column=5, value=lbl2).font = Font(bold=True, size=10)
            ws.cell(row=r, column=5).fill = lbl_f
            ws.cell(row=r, column=5).border = tb
            ws.cell(row=r, column=6, value=val2).border = tb
        else:
            ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
            ws.cell(row=r, column=4, value=val).border = tb
        ws.row_dimensions[r].height = 17

    ws.row_dimensions[7].height = 8

    ws.merge_cells(start_row=8, start_column=2, end_row=12, end_column=2)
    ws["B8"].value = "폐기물품"
    ws["B8"].alignment = C
    ws["B8"].font = sec_ft
    ws["B8"].fill = sec_f
    ws["B8"].border = tb

    for ci, hdr in [(3, "물  품"), (4, "수  량"), (5, "기타사항")]:
        c = ws.cell(row=8, column=ci, value=hdr)
        c.font = sec_ft; c.fill = sec_f; c.alignment = C; c.border = tb
    ws.merge_cells("E8:F8")

    disp_qtys = _agg_disposal(app.assets)
    for idx, name in enumerate(_DISPOSAL_ROWS):
        r = 9 + idx
        qty = disp_qtys[name]
        ws.cell(row=r, column=3, value=f"  {name}").border = tb
        ws.cell(row=r, column=3).font = Font(size=10)
        ws.cell(row=r, column=4, value=qty if qty else "").border = tb
        ws.cell(row=r, column=4).alignment = C
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)
        ws.cell(row=r, column=5).border = tb
        ws.row_dimensions[r].height = 16

    ws.row_dimensions[14].height = 8

    ws.merge_cells(start_row=15, start_column=2, end_row=18, end_column=2)
    ws["B15"].value = "폐기\n사유등"
    ws["B15"].alignment = C
    ws["B15"].font = sec_ft
    ws["B15"].fill = sec_f
    ws["B15"].border = tb

    reason_rows = [
        ("폐기사유 -", "노후 및 기능 저하로 인한 폐기"),
        ("폐기일자 -", wr.wipe_date if wr else ""),
        ("폐기장소 -", wr.wipe_location if wr else ""),
        ("폐기방법 -", wr.destruction_method if wr else ""),
    ]
    for idx, (lbl, val) in enumerate(reason_rows):
        r = 15 + idx
        ws.cell(row=r, column=3, value=f"  {lbl}").border = tb
        ws.cell(row=r, column=3).font = Font(size=10)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
        ws.cell(row=r, column=4, value=val).border = tb
        ws.cell(row=r, column=4).font = Font(size=10)
        ws.row_dimensions[r].height = 17

    ws.row_dimensions[19].height = 10
    ws.merge_cells("B20:F20")
    ws["B20"].value = "「폐기물관리법」 제13조에 따라 위와 같이 결함제품을 폐기하였음을 확인합니다."
    ws["B20"].font = Font(size=10)
    ws["B20"].alignment = C
    ws.row_dimensions[20].height = 18
    ws.row_dimensions[21].height = 10

    ws.merge_cells("D22:F22")
    ws["D22"].value = "년     월     일"
    ws["D22"].font = Font(size=11)
    ws["D22"].alignment = C
    ws.row_dimensions[22].height = 20
    ws.row_dimensions[23].height = 10

    ws.merge_cells("C24:F24")
    ws["C24"].value = f"폐기처리자  {wr.technician_name if wr else ''}  ({wr.technician_company if wr else ''})                 (서명)"
    ws["C24"].font = Font(size=11)
    ws["C24"].alignment = L
    ws.row_dimensions[24].height = 22

    try:
        stamp_bytes = base64.b64decode(STAMP_B64)
        stamp3 = XLImage(BytesIO(stamp_bytes))
        stamp3.width, stamp3.height = 80, 80
        ws.add_image(stamp3, "E22")
    except Exception as e:
        print(f"[직인] 폐기물품확인서 삽입 실패: {e}")


def _generate_disposal_excel(app) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "폐기물품확인서"
    _build_disposal_sheet(ws, app)
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def _generate_destruction_excel(app) -> BytesIO:
    """파쇄증명서 + 물품인수인계증 + 폐기물품확인서 3종 합본"""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "파쇄증명서"
    _build_shredding_sheet(ws1, app)
    ws2 = wb.create_sheet("물품인수인계증")
    _build_transfer_sheet(ws2, app)
    ws3 = wb.create_sheet("폐기물품확인서")
    _build_disposal_sheet(ws3, app)
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


@router.get("/applications/{app_id}/report/blancco")
def download_blancco_report(request: Request, app_id: int, db: Session = Depends(get_db)):
    user, redir = _check(request)
    if redir:
        return redir

    app = db.query(models.Application).filter(models.Application.id == app_id).first()
    if not app or not app.data_wipe_record:
        return RedirectResponse(f"/admin/applications/{app_id}", status_code=302)

    wr = app.data_wipe_record

    # 증명서가 업로드된 경우 증명서 우선 제공
    if wr.blancco_certificate_file and os.path.exists(wr.blancco_certificate_file):
        fpath = Path(wr.blancco_certificate_file)
        ext = fpath.suffix.lower()
        media_type = "application/pdf" if ext == ".pdf" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if ext in (".xlsx", ".xls") else "application/octet-stream"
        fname = f"blancco_certificate_{app_id}{ext}"
        return FileResponse(path=str(fpath), media_type=media_type, filename=fname)

    if not wr.blancco_report_file or not os.path.exists(wr.blancco_report_file):
        return RedirectResponse(f"/admin/applications/{app_id}?error=no_blancco_file", status_code=302)

    fpath = Path(wr.blancco_report_file)
    ext = fpath.suffix.lower()
    media_type = "application/pdf" if ext == ".pdf" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if ext in (".xlsx", ".xls") else "application/octet-stream"
    fname = f"blancco_report_{app_id}{ext}"
    return FileResponse(path=str(fpath), media_type=media_type, filename=fname)


@router.post("/applications/{app_id}/blancco/receive")
def blancco_receive(request: Request, app_id: int, db: Session = Depends(get_db)):
    """블랑코 영구삭제 보고서 수령완료 처리"""
    user, redir = _check(request)
    if redir:
        return redir
    if user["role"] != "coretail":
        return RedirectResponse(f"/admin/applications/{app_id}", status_code=302)

    app = db.query(models.Application).filter(models.Application.id == app_id).first()
    if app and app.data_wipe_record:
        wr = app.data_wipe_record
        wr.blancco_received = True
        wr.blancco_received_at = datetime.now()
        wr.updated_at = datetime.now()
        db.commit()

    return RedirectResponse(f"/admin/applications/{app_id}", status_code=302)


@router.post("/applications/{app_id}/blancco/upload_cert")
async def blancco_upload_cert(request: Request, app_id: int, db: Session = Depends(get_db)):
    """블랑코 영구삭제 증명서 업로드"""
    user, redir = _check(request)
    if redir:
        return redir
    if user["role"] != "coretail":
        return RedirectResponse(f"/admin/applications/{app_id}", status_code=302)

    form = await request.form()
    app = db.query(models.Application).filter(models.Application.id == app_id).first()
    if not app or not app.data_wipe_record:
        return RedirectResponse(f"/admin/applications/{app_id}", status_code=302)

    wr = app.data_wipe_record
    cert_file = form.get("blancco_certificate")
    if cert_file and getattr(cert_file, "filename", None):
        cert_dir = Path("static/uploads/blancco_certificates")
        cert_dir.mkdir(parents=True, exist_ok=True)
        ext = Path(cert_file.filename).suffix.lower()
        if ext not in (".pdf", ".xlsx", ".xls"):
            ext = ".pdf"
        fname = f"{app_id}_blancco_cert{ext}"
        fpath = cert_dir / fname
        contents = await cert_file.read()
        if contents:
            fpath.write_bytes(contents)
            wr.blancco_certificate_file = str(fpath).replace("\\", "/")
            wr.updated_at = datetime.now()
            db.commit()

    return RedirectResponse(f"/admin/applications/{app_id}", status_code=302)


@router.get("/applications/{app_id}/report/shredding")
def download_shredding_report(request: Request, app_id: int, db: Session = Depends(get_db)):
    """파쇄증명서 다운로드"""
    user, redir = _check(request)
    if redir:
        return redir

    app = db.query(models.Application).filter(models.Application.id == app_id).first()
    if not app or not app.data_wipe_record:
        return RedirectResponse(f"/admin/applications/{app_id}", status_code=302)

    output = _generate_shredding_excel(app)
    fname = f"파쇄증명서_{app_id}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{fname}"},
    )


@router.get("/applications/{app_id}/report/transfer")
def download_transfer_report(request: Request, app_id: int, db: Session = Depends(get_db)):
    """물품인수인계증 다운로드"""
    user, redir = _check(request)
    if redir:
        return redir

    app = db.query(models.Application).filter(models.Application.id == app_id).first()
    if not app or not app.data_wipe_record:
        return RedirectResponse(f"/admin/applications/{app_id}", status_code=302)

    output = _generate_transfer_excel(app)
    fname = f"물품인수인계증_{app_id}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{fname}"},
    )


@router.get("/applications/{app_id}/report/disposal")
def download_disposal_report(request: Request, app_id: int, db: Session = Depends(get_db)):
    """폐기물품확인서 다운로드"""
    user, redir = _check(request)
    if redir:
        return redir

    app = db.query(models.Application).filter(models.Application.id == app_id).first()
    if not app or not app.data_wipe_record:
        return RedirectResponse(f"/admin/applications/{app_id}", status_code=302)

    output = _generate_disposal_excel(app)
    fname = f"폐기물품확인서_{app_id}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{fname}"},
    )


@router.post("/applications/{app_id}/pricing")
async def set_pricing(request: Request, app_id: int, db: Session = Depends(get_db)):
    user, redir = _check(request)
    if redir:
        return redir
    if user["role"] != "coretail":
        return RedirectResponse(f"/admin/applications/{app_id}", status_code=302)

    form = await request.form()

    app = db.query(models.Application).filter(
        models.Application.id == app_id,
        models.Application.status == "collected",
    ).first()

    if app:
        total = 0.0
        for asset in app.assets:
            key = f"price_{asset.id}"
            try:
                price = float(form.get(key, 0) or 0)
            except ValueError:
                price = 0.0
            asset.unit_price = price
            total += price * asset.quantity

        if not app.settlement:
            settlement = models.Settlement(application_id=app_id)
            db.add(settlement)
            db.flush()
            db.refresh(app)

        fee_rate = _get_fee_rate(db)
        branch_total = total * (1 - fee_rate / 100)

        app.settlement.total_amount = total
        app.settlement.welfare_fee_rate = fee_rate
        app.settlement.branch_total_amount = branch_total
        app.settlement.pricing_notes = form.get("pricing_notes", "")
        app.status = "priced"
        app.updated_at = datetime.now()
        db.commit()

    return RedirectResponse(f"/admin/applications/{app_id}", status_code=302)


@router.get("/settings", response_class=HTMLResponse)
def settings_page(request: Request, db: Session = Depends(get_db)):
    user, redir = _check(request)
    if redir:
        return redir
    if user["role"] != "coretail":
        return RedirectResponse("/admin/dashboard", status_code=302)

    fee_rate = _get_fee_rate(db)
    return templates.TemplateResponse(
        "admin/settings.html",
        {"request": request, "session": request.session, "welfare_fee_rate": fee_rate},
    )


@router.post("/settings")
async def save_settings(request: Request, db: Session = Depends(get_db)):
    user, redir = _check(request)
    if redir:
        return redir
    if user["role"] != "coretail":
        return RedirectResponse("/admin/dashboard", status_code=302)

    form = await request.form()
    try:
        rate = float(form.get("welfare_fee_rate", 0) or 0)
        rate = max(0.0, min(100.0, rate))
    except (ValueError, TypeError):
        rate = 0.0

    config = db.query(models.SystemConfig).filter(models.SystemConfig.key == "welfare_fee_rate").first()
    if config:
        config.value = str(rate)
        config.updated_at = datetime.now()
    else:
        db.add(models.SystemConfig(key="welfare_fee_rate", value=str(rate)))
    db.commit()

    return RedirectResponse("/admin/settings", status_code=302)


@router.post("/applications/{app_id}/complete")
def complete_payment(request: Request, app_id: int, db: Session = Depends(get_db)):
    user, redir = _check(request)
    if redir:
        return redir
    if user["role"] != "welfare":
        return RedirectResponse(f"/admin/applications/{app_id}", status_code=302)

    app = db.query(models.Application).filter(
        models.Application.id == app_id,
        models.Application.status == "branch_confirmed",
    ).first()

    if app and app.settlement:
        app.settlement.welfare_confirmed = True
        app.settlement.welfare_confirmed_at = datetime.now()
        app.settlement.payment_confirmed = True
        app.settlement.payment_date = datetime.now().strftime("%Y-%m-%d")
        app.status = "completed"
        app.updated_at = datetime.now()
        db.commit()

    return RedirectResponse(f"/admin/applications/{app_id}", status_code=302)
