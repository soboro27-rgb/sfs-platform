from fastapi import APIRouter, Request, Depends, UploadFile, File
from fastapi.responses import RedirectResponse, HTMLResponse, StreamingResponse
from sqlalchemy.orm import Session
from database import get_db
import models
from auth import require_admin, hash_password
from config import templates
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

router = APIRouter()

ROLE_LABEL = {"branch": "지점 담당자", "welfare": "주관사", "coretail": "운영사(코어테일)"}


def _check(request: Request):
    user = require_admin(request)
    if not user:
        return None, RedirectResponse("/login", status_code=302)
    return user, None


@router.get("/users", response_class=HTMLResponse)
def user_list(request: Request, role: str = "", db: Session = Depends(get_db)):
    user, redir = _check(request)
    if redir:
        return redir

    query = db.query(models.User)
    if role:
        query = query.filter(models.User.role == role)
    users = query.order_by(models.User.role, models.User.branch_code).all()

    return templates.TemplateResponse("admin/users.html", {
        "request": request,
        "session": request.session,
        "users": users,
        "current_role": role,
        "ROLE_LABEL": ROLE_LABEL,
    })


@router.get("/users/new", response_class=HTMLResponse)
def new_user_page(request: Request):
    user, redir = _check(request)
    if redir:
        return redir
    return templates.TemplateResponse("admin/user_form.html", {
        "request": request,
        "session": request.session,
        "error": None,
        "ROLE_LABEL": ROLE_LABEL,
    })


@router.post("/users/new")
async def create_user(request: Request, db: Session = Depends(get_db)):
    user, redir = _check(request)
    if redir:
        return redir

    form = await request.form()
    branch_code = form.get("branch_code", "").strip()
    password = form.get("password", "").strip()
    branch_name = form.get("branch_name", "").strip()
    region = form.get("region", "").strip()
    role = form.get("role", "branch")

    # welfare는 branch만 생성 가능
    if role in ("welfare", "coretail") and user["role"] != "coretail":
        role = "branch"

    if not branch_code or not password or not branch_name:
        return templates.TemplateResponse("admin/user_form.html", {
            "request": request,
            "session": request.session,
            "error": "필수 항목을 모두 입력해주세요.",
            "ROLE_LABEL": ROLE_LABEL,
        })

    existing = db.query(models.User).filter(models.User.branch_code == branch_code).first()
    if existing:
        return templates.TemplateResponse("admin/user_form.html", {
            "request": request,
            "session": request.session,
            "error": f"이미 존재하는 아이디입니다: {branch_code}",
            "ROLE_LABEL": ROLE_LABEL,
        })

    db.add(models.User(
        branch_code=branch_code,
        password_hash=hash_password(password),
        branch_name=branch_name,
        region=region,
        role=role,
        created_at=datetime.now(),
    ))
    db.commit()
    return RedirectResponse("/admin/users", status_code=302)


@router.post("/users/{user_id}/reset-password")
async def reset_password(request: Request, user_id: int, db: Session = Depends(get_db)):
    current_user, redir = _check(request)
    if redir:
        return redir
    if current_user["role"] != "coretail":
        return RedirectResponse("/admin/users", status_code=302)

    form = await request.form()
    new_password = form.get("new_password", "").strip()
    if not new_password:
        return RedirectResponse("/admin/users", status_code=302)

    target = db.query(models.User).filter(models.User.id == user_id).first()
    if target:
        target.password_hash = hash_password(new_password)
        db.commit()
    return RedirectResponse("/admin/users", status_code=302)


@router.post("/users/{user_id}/toggle")
def toggle_user(request: Request, user_id: int, db: Session = Depends(get_db)):
    current_user, redir = _check(request)
    if redir:
        return redir

    target = db.query(models.User).filter(models.User.id == user_id).first()
    if target and target.id != current_user["user_id"]:
        target.is_active = not target.is_active
        db.commit()
    return RedirectResponse("/admin/users", status_code=302)


@router.get("/users/template")
def download_template(request: Request):
    user, redir = _check(request)
    if redir:
        return redir

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "계정목록"

    header_fill = PatternFill(start_color="005B30", end_color="005B30", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    headers = ["지점코드(아이디)*", "비밀번호*", "지점명/이름*", "지역", "역할*"]
    col_widths = [22, 18, 24, 14, 28]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22

    example_fill = PatternFill(start_color="F0F7F2", end_color="F0F7F2", fill_type="solid")
    examples = [
        ["MG011", "Branch!2024", "서울중구지점", "서울", "branch"],
        ["MG012", "Branch!2024", "부산중구지점", "부산", "branch"],
        ["WELFARE02", "Welfare!2024", "주관사관리자2", "전국", "welfare"],
    ]
    for row_idx, row_data in enumerate(examples, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = example_fill

    # 작성요령 시트
    ws2 = wb.create_sheet("작성요령")
    ws2["A1"] = "작성 요령"
    ws2["A1"].font = Font(bold=True, size=13)
    notes = [
        ("지점코드(아이디)", "로그인 ID. 영문+숫자 조합 권장. 중복 불가."),
        ("비밀번호", "초기 비밀번호. 영문+숫자+특수문자 조합 권장."),
        ("지점명/이름", "지점명 또는 담당자/기관 이름."),
        ("지역", "지역명 (서울, 부산 등). 비워도 됨."),
        ("역할", "branch = 지점 담당자 / welfare = 주관사 관리자 / coretail = 운영사 관리자"),
    ]
    for i, (field, desc) in enumerate(notes, 3):
        ws2[f"A{i}"] = field
        ws2[f"A{i}"].font = Font(bold=True)
        ws2[f"B{i}"] = desc
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 60

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''%EA%B3%84%EC%A0%95%EB%93%B1%EB%A1%9D%EC%96%91%EC%8B%9D.xlsx"},
    )


@router.get("/users/bulk", response_class=HTMLResponse)
def bulk_upload_page(request: Request):
    user, redir = _check(request)
    if redir:
        return redir
    return templates.TemplateResponse("admin/bulk_upload.html", {
        "request": request,
        "session": request.session,
        "results": None,
    })


@router.post("/users/bulk")
async def bulk_upload(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    user, redir = _check(request)
    if redir:
        return redir

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(BytesIO(contents))
    except Exception:
        return templates.TemplateResponse("admin/bulk_upload.html", {
            "request": request,
            "session": request.session,
            "results": {"error": "올바른 엑셀 파일(.xlsx)이 아닙니다."},
        })

    ws = wb.active
    success, failed = [], []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue

        branch_code = str(row[0]).strip() if row[0] else ""
        password = str(row[1]).strip() if row[1] else ""
        branch_name = str(row[2]).strip() if row[2] else ""
        region = str(row[3]).strip() if row[3] else ""
        role = str(row[4]).strip().lower() if row[4] else "branch"

        if not branch_code or not password or not branch_name:
            failed.append({"code": branch_code or "(빈칸)", "reason": "필수 항목 누락"})
            continue

        if role not in ("branch", "welfare", "coretail"):
            role = "branch"

        if role in ("welfare", "coretail") and user["role"] != "coretail":
            failed.append({"code": branch_code, "reason": "권한 없음 (welfare/coretail은 코어테일만 생성 가능)"})
            continue

        existing = db.query(models.User).filter(models.User.branch_code == branch_code).first()
        if existing:
            failed.append({"code": branch_code, "reason": "중복 아이디"})
            continue

        try:
            db.add(models.User(
                branch_code=branch_code,
                password_hash=hash_password(password),
                branch_name=branch_name,
                region=region,
                role=role,
                created_at=datetime.now(),
            ))
            db.flush()
            success.append({"code": branch_code, "name": branch_name, "role": role})
        except Exception as e:
            failed.append({"code": branch_code, "reason": str(e)})

    db.commit()

    return templates.TemplateResponse("admin/bulk_upload.html", {
        "request": request,
        "session": request.session,
        "results": {"success": success, "failed": failed},
    })
